
#https://medium.com/analytics-vidhya/gui-for-your-python-program-with-multiple-windows-options-78c2ea8d259d

#import openpyxl and tkinter modules
import matplotlib.axes
#matplotlib.axes.AxesSubplot
import matplotlib
import matplotlib.pyplot as plt
import calendar
from babel.dates import format_date, parse_date, get_day_names, get_month_names
from babel.numbers import *  # Additional Import```
import babel.numbers

from re import match as re_match

from tkcalendar import Calendar,DateEntry

from datetime import date
import pandas as pd
from pandas import Series, DataFrame
from datetime import datetime
import time
import math
from tkinter.ttk import *
from tkinter import *
from tkinter import messagebox
from openpyxl import *
import numpy_financial as npf
import numpy as np
import xlsxwriter
import re
import fpdf
from win32com import client
import win32api
import seaborn as sns
import os
import os.path
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string
import win32com.client
from pywintypes import com_error
from openpyxl.chart import (
    LineChart,
    Reference,
)

from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart import (
    DoughnutChart,
    Reference,
    Series,
)
from openpyxl.chart.label import DataLabelList
import win32com.client



#old
import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showinfo
from tkinter import Tk, Label, Button, StringVar
from tkinter import messagebox



class Financial_Calculator(tk.Tk):


    def __init__(self,*args, **kwargs):
        #options()

        tk.Tk.__init__(self, *args, **kwargs)
        self.wm_title('Financial Calculators')

        container = tk.Frame(self, height = 400, width = 600)
        
        tk.Tk.geometry(self,'800x350')
        container.pack(side = 'top', fill = 'both', expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frame = {}

        for F in (StartPage,FutureValue,NPV,Mortgage,EffectiveInterest):
            frame = F(container, self)

            self.frame[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)


    def show_frame(self, cont):

        frame = self.frame[cont]
        frame.tkraise()


class StartPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text = "CALCULATOR MENU")
        #label.pack(pady=10, padx=10)
        label.grid(row=1,column = 0, sticky ='news')

        #button1 = ttk.Button(self, text = "Future Value",command = lambda: controller.show_frame(FutureValue))
        Button(self,bg = 'orange', text = "Future Value",command = lambda: controller.show_frame(FutureValue)).grid(row=2, column = 0, sticky ='news')
        #button2 = ttk.Button(self, text = "NPV & IRR",command = lambda: controller.show_frame(NPV))
        Button(self,bg = 'orange', text = "NPV & IRR",command = lambda: controller.show_frame(NPV)).grid(row=3, column = 0, sticky ='news')
        
        #button3 = ttk.Button(self, text = "Mortgage",command = lambda: controller.show_frame(Mortgage))
        Button(self, bg = 'orange', text = "Mortgage",command = lambda: controller.show_frame(Mortgage)).grid(row=4, column = 0, sticky ='news')
        #button4 = ttk.Button(self, text = "Effective Interest",command = lambda: controller.show_frame(EffectiveInterest))
        Button(self, bg = 'orange', text = "Effective Interest",command = lambda: controller.show_frame(EffectiveInterest)).grid(row=5, column = 0, sticky ='news')
        #button5 = ttk.Button(self, text = "Exit",command = lambda: controller.destroy())
        Button(self, bg = 'red', text = "Exit",command = lambda: controller.destroy()).grid(row=6, column = 0, sticky ='news')

        self.columnconfigure(0, weight=1)  # column on left
        #self.columnconfigure(1, weight=1)  # column on right
        self.rowconfigure(0, weight=1)     # row above
        self.rowconfigure(1, weight=1)     # row below 
    
        #button1.pack()
        #Button(self, bg = 'orange').grid(row=1, column = 5, sticky ='news',)
        #button2.pack()
        #button3.pack()
        #button4.pack()
        #button5.pack()

    #def exit(self):
        #self.destroy()

class FutureValue(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        Label(self, text = "FUTURE VALUE").grid(row=0,column=1)
    
        Label(self,text = "Starting Amount:").grid(row=1)

        Label(self, text = "Yearly Deposit:").grid(row=2)

        Label(self, text = "Total Returns:",font=('calibri',10, 'bold')).grid (row=1, column=3)
        Label(self, text = "Total Interest:",font=('calibri',10, 'bold')).grid (row=2, column=3)
        Label(self, text = "Total Deposit:",font=('calibri',10, 'bold')).grid (row=3, column=3)
        Label(self, text = "Present Value:",font=('calibri',10, 'bold')).grid (row=4, column=3)

        Label(self, text = "Annual Interest:",justify="right").grid(row=3)
        Label(self, text = "No of Years:",justify="right").grid (row=4)
        Label(self, text = "Yearly Deposit: at Beg enter(1), at End enter (0)",wraplength=90, justify="right").grid (row=5)

        Label(self, text = "Developer : Sunil",font=('calibri',10, 'bold', 'italic')).place(x =480, y = 300)

    #Entry box for enter details
        prin = tk.Entry(self)
        pmt = tk.Entry(self)
        interest = tk.Entry(self)
        n = tk.Entry(self)
        ant = tk.Entry(self)

        #Entry field location
        prin.grid(row=1,column=1)
        pmt.grid(row=2,column=1)
        interest.grid(row=3,column=1)
        n.grid(row=4, column=1)
        ant.grid(row=5, column=1)

    #Result output field format

        blank_fv = tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_fv.grid(row=1, column=4)

        blank_int = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_int.grid(row=2, column=4)

        blank_prin = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_prin.grid(row=3, column=4)

        blank_pvt = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_pvt.grid(row=4, column=4)

    #Quit application

        def quit():
        
            MsgBox = messagebox.askquestion ('Exit Application','Are you sure you want to "Exit"',icon = 'warning')
            if MsgBox == 'yes':
                self.destroy()
            else:
                messagebox.showinfo('Return','You will now return to the application')

    # Clear results fields
        def clearfv():
            blank_fv.delete(0, END)
            blank_int.delete(0, END)
            blank_prin.delete(0, END)
            blank_pvt.delete(0, END)
    #Focus to strating amount
            #prin.focus_set()

    #Clear all fields
        def clearall():
            blank_fv.delete(0, END)
            prin.delete(0, END)
            interest.delete(0, END)
            n.delete(0, END)
            pmt.delete(0, END)
            ant.delete(0, END)
            blank_int.delete(0, END)
            blank_prin.delete(0, END)
            blank_pvt.delete(0, END)
            
    #Set curser focus on to strating amount
            prin.focus_set()

    #Calculations

        def show_answer():
    
    #Setting variables, data type from entry
    
    #Initial deposit
            prin_x1 = prin.get()
            if prin.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Starting Amount"')
            try:
                int(prin_x1)
            except ValueError:
                try:
                    float(prin_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            prin_x =float(prin_x1)

    #Yearly deposit
            pmt_x1 = pmt.get()
            if pmt.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Yearly Deposit"')
            try:
                int(pmt_x1)
            except ValueError:
                try:
                    float(pmt_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            pmt_x =float(pmt_x1) 

    #Interest
            rate_x =interest.get()
            if interest.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Annual Interest"')
            try:
                int(rate_x)
            except ValueError:
                try:
                    float(rate_x)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            rate = float(rate_x)
        
    #years
            yrs_x = n.get()
            if n.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "No of Years"')
            try:
                int(yrs_x)
            except ValueError:
                try:
                    float(yrs_x)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            yrs = float(yrs_x)
     
    #Yearly deposit at (Big/End)of each year and data type
            b_e_x = ant.get()
            if ant.get() == "":
                return messagebox.showwarning('Warning', 'Please enter 1 for Beg or 0 for End')
            elif not re.match("^[01]*$", ant.get()):
                return messagebox.showwarning('Warning', 'Please enter 1 for Beg or 0 for End')
            b_e = int(b_e_x)
            
          
    #Interst calculation factor
            i= (1+(rate/100))**yrs
            #print(i)
    #Interest in percentage
            ir = float(rate/100)
    #Total amount for initial amount with interest
            prin_p = float(prin_x) * float(i)
    #Total amount for pmt (pay at begining of the year) amount with interest    
            (pmt_p_b) = float(pmt_x)* ((i-1)/(rate/100)) * (1 + (rate/100))
    #Total amount for pmt (pay at end of the year) amount with interest       
            (pmt_p_e) = float(pmt_x)* ((i-1)/(rate/100))
            #print(pmt_p)
    #Total feature value with interest
            if b_e == 1:
                fv = prin_p + pmt_p_b
            else:
                fv = prin_p + pmt_p_e
            Ans = ('${:,.2f}'.format(fv))
            
            #Ans = format(fv,'.2f')
            blank_fv.insert(3,Ans) 
                                                        
    #Total interest accumulated
            total_intst = fv - (float(prin_x) + float(pmt_x)*yrs)
            #print(total_interest)
            Ans1=('${:,.2f}'.format(total_intst))
            #Ans1 = format(total_intst,'.2f')
            blank_int.insert(3,Ans1)

    #Present Value Calculations
            #pvt = ttr*( 1 / (1+ (rate/100))**ttyrs)
            pvt = fv*( 1 / (1+ (rate/100))**yrs)
            Ans3 =('${:,.2f}'.format(pvt))
            blank_pvt.insert(3,Ans3)

        
    #Total amount deposited including principle and pmt
            total_prin = fv - total_intst
            #print(total_prin)
            Ans2 = ('${:,.2f}'.format(total_prin))
            #Ans2 = format(total_prin,'.2f')
            
            blank_prin.insert(3,Ans2)
            return messagebox.showinfo('Created', 'See Results')

    #Set curser focus on to strating amount
            prin.focus_set()



    #To create excel schedule
        def show_schedule():

    #Total returns is empty, show message
            
            if blank_fv.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
                blank_fv.get()
            
    #Total interest is empty, show message
            
            if blank_int.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
               blank_int.get()
            
    #Total deposit is empty, show message
            
            if blank_prin.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
                blank_prin.get()
        
    #Show Schedule
    #Declaring variable again and data type
            prin_x1 = prin.get()
            prin_x =float(prin_x1)

    #Yearly periodical amount
            pmt_x1 = pmt.get()
            pmt_x =float(pmt_x1) 
        
    #Expected interest rate
            rate_x =interest.get()
            rate = float(rate_x)
        
    #Total number of years
            yrs_x = n.get()
            yrs = float(yrs_x)
     
    #Anuity option, yearly deposit as Big or End
            b_e_x = ant.get()
            b_e = int(b_e_x)
    #interest in percentage
            ir = float(rate/100)
            
    #Shecdule calculations
    #No of years
            
            period = []
            for i in range (1, int(yrs + 1)):
                period.append(i)
                #print (str(i))
            
    #Start principal
            sp_list = []
            for i in range (1, int(yrs +1)):
                if b_e ==1:
                    sp = prin_x + (pmt_x * i)
                else:
                    sp = prin_x + (pmt_x * i)-pmt_x

                sp_list.append(sp)
                #print (str(round(sp, 2)))
            sp_list_np = np.array(sp_list)
            #print(sp_list_np)
        

    #End balance
            eb_list = []
            for i in range (1, int(yrs +1)):

                if b_e ==1:

                    pv_f = prin_x*(( int(1)+ir)**i)
                    pmt_f = pmt_x *(int(1) + ir* b_e)/ir*((int(1) + ir)**i -int(1))
                    eb = pv_f + pmt_f

                else:
                    pv_f = prin_x*(( int(1)+ir)**i)
                    pmt_f = pmt_x *(int(1) + ir* b_e)/ir*((int(1) + ir)**i -int(1))
                    eb = pv_f + pmt_f
                
                eb_list.append(eb)
                #print (str(round(eb, 2)))
            eb_list_np = np.array(eb_list)
            #print(eb_list_np)

    #Cumulative interest
            ci_list = []
            for i in range (1, int(yrs +1)):
                if b_e ==1:
                    pv_f = prin_x*(( int(1)+ir)**i)
                    pmt_f = pmt_x *(int(1) + ir* b_e)/ir*((int(1) + ir)**i -int(1))
                    eb = pv_f + pmt_f
                
                    sp = prin_x + (pmt_x * i)
                    ci = (eb-sp)
                else:
                    pv_f = prin_x*(( int(1)+ir)**i)
                    pmt_f = pmt_x *(int(1) + ir* b_e)/ir*((int(1) + ir)**i -int(1))
                    eb = pv_f + pmt_f
                
                
                    ep = prin_x + (pmt_x * i)
                
                    ci = (eb-ep)
            
                ci_list.append(ci)
                #print (str(round(ci, 2)))
            ci_list_np = np.array(ci_list)
            #print(ci_list_np)

    #End principal
            ep_list = []
            for i in range (1, int(yrs +1)):
                if b_e ==1:
                    ep = prin_x + (pmt_x * i)
                else:
                    ep = prin_x + (pmt_x * i)

                ep_list.append(ep)
                #print (str(round(ep, 2)))
            ep_list_np = np.array(ep_list)
            #print(ep_list_np)
        

    #Start balance(not added yet)
            sb_list = []
            for i in range (1, int((yrs) +1)):

                if b_e ==1:

                    pv_f = prin_x*(( int(1)+ir)**i)
                    pmt_f = pmt_x *(int(1) + ir* b_e)/ir*((int(1) + ir)**i -int(1))
                    sb = pv_f + pmt_f

                else:
                    pv_f = prin_x*(( int(1)+ir)**i)
                    pmt_f = pmt_x *(int(1) + ir* b_e)/ir*((int(1) + ir)**i -int(1))
                    sb = pv_f + pmt_f
                
                sb_list.append(sb)
                #print (str(round(sb, 2)))
            sb_list_np = np.array(sb_list)
            #print(sb_list_np)


    #Variable for gragh and Chart

            ttrr = blank_fv.get()
            value = re.sub("[^\d\.]", "",ttrr )#make numbers only
            ttr = float(value)
            #print(ttr)
            
            ttid = float(prin.get())
            ttpmt = float(pmt.get())
            ttyrs = float(n.get())
            ttpd = float(ttpmt*ttyrs)
            
            ttii = blank_int.get()
            value = re.sub("[^\d\.]", "",ttii)#make numbers only
            tti = float(value)
            #print(tti)            
            
            rate = float(interest.get())
            irs = rate/100
            
    #Present Value Calculations
            pvt = ttr*( 1 / (1+ (rate/100))**ttyrs)
            #print(pvt)
            
    #Export to excel

            workbook = xlsxwriter.Workbook(r'C:\\temp\\Schedule.xlsx')
            worksheet = workbook.add_worksheet('Schedule')
            

    # Create a dictionary with your file names and your files
            all_files = {'Years': period,
                        'Starting Deposit': sp_list,
                        'Ending Deposit' : ep_list,
                        'Interest': ci_list,
                        'Yearly Balance': eb_list,
                        }

            bold = workbook.add_format({'bold': True})
            cell_format = workbook.add_format({'bold': True, 'font_color': 'black','align': 'right'})
            cell_format_a = workbook.add_format({'bold': True, 'font_color': 'black','align': 'left'})
            cell_format.set_font_size(12)
            cell_format.set_align('right')
            
            cell_format1 = workbook.add_format({'font_color': 'black'})
            cell_format1.set_font_size(12)
            cell_format1.set_align('right')
            

            cell_format2 = workbook.add_format({'font_color': 'black', 'bold': True})
            cell_format2.set_font_size(12)
            cell_format2.set_align('right')

            cell_format3 = workbook.add_format({'font_color': 'black', 'bold': True,'num_format': '$#,##0.00'})
            cell_format3.set_font_size(13)
            cell_format3.set_align('right') 

            format = workbook.add_format({'num_format': '$#,##0.00'})
            num2 = workbook.add_format({'num_format': '#,##0.00%'})
            num3 = workbook.add_format({'num_format': '#,##0'})
            num4 = workbook.add_format({'num_format': 'mmm, d yyyy hh:mm AM/PM'})
            

            
            worksheet.set_column('B1:E1',1,cell_format1)
            worksheet.set_column('F1:H1',2,cell_format1)
            worksheet.set_column('A1:A1',1,cell_format1)
            worksheet.write('G1','Developer:Sunil', cell_format_a)
            worksheet.write('G2','Total Returns', cell_format_a)
            worksheet.write('G3','Present Value', cell_format_a)
            worksheet.write('G4','Starting Amount', cell_format_a)
            worksheet.write('G5','Yearly Deposit', cell_format_a)
            worksheet.write('G6','Total Interest', cell_format_a)
            worksheet.write('G7','No of Years', cell_format_a)
            worksheet.write('G8','Interest Rate', cell_format_a)
            
            worksheet.write('H2',ttr, cell_format3)
            worksheet.write('H3',pvt, cell_format3)
            worksheet.write('H4',ttid, cell_format3)
            worksheet.write('H5',ttpd, cell_format3)
            worksheet.write('H6',tti, cell_format3)
            worksheet.write('H7',ttyrs, cell_format1)
            worksheet.write('H8',irs, num2)
            

            for index, file in enumerate(all_files):
                worksheet.write(0,index, file, cell_format)
                for row in range(1,len(all_files[file])+1):
                    worksheet.write(row, index, all_files[file][row-1])
                    worksheet.set_column('B:E',1,format)
                    
    # Formatted as 'dd/mm/yy'
            date_time = datetime.now()
            worksheet.write_datetime('H1', date_time,num4)
            
            #workbook.save(r'C:\\temp\\Schedule.xlsx')
            workbook.close()
            
    #Set_column_width auto(worksheet):

            workbook = load_workbook(r'C:\\temp\\Schedule.xlsx')
            worksheet = workbook['Schedule']

            for col in worksheet.columns:
                 max_length = 0
                 column = col[0].column_letter # Get the column name
                 for cell in col:
                     try: # Necessary to avoid error on empty cells
                         if len(str(cell.value)) > max_length:
                             max_length = len(str(cell.value))
                     except:
                         pass
                 adjusted_width = (max_length + 2) * 1.2
                 worksheet.column_dimensions[column].width = adjusted_width

            workbook.save(r'C:\\temp\\Schedule.xlsx')
            workbook.close() 



    #Generating line chart in excel
            workbook = load_workbook(r'C:\\temp\\Schedule.xlsx')
            worksheet = workbook['Schedule']

    #Cell references (original spreadsheet) 
            min_column = workbook.active.min_column
            max_column = workbook.active.max_column
            min_row = workbook.active.min_row
            max_row = workbook.active.max_row

    #Line chart
            c1 = LineChart()
    #locate data and categories
            data = Reference(worksheet,
                     min_col=min_column+1,
                     #min_col=min_column+3,      
                     max_col=max_column-3,
                     min_row=min_row,
                     max_row=max_row) #including headers
            categories = Reference(worksheet,
                           min_col=min_column,
                           max_col=min_column,
                           min_row=min_row+1,
                           max_row=max_row) #not including headers
    #Adding data and categories
            c1.add_data(data, titles_from_data=True)
            c1.set_categories(categories)
    #Location chart
            worksheet.add_chart(c1, "F10")
            c1.title = 'Balance Accumulation Graph'
            c1.y_axis.title = 'Value'
            c1.x_axis.title = 'Period'
            c1.style = 26 #choose the chart style
            workbook.save(r'C:\\temp\\Schedule.xlsx')
            #return messagebox.showinfo('Schedule.xlsx created', 'Please check c:\\ temp folder')

    #Generating Pie chart
            workbook = load_workbook(r'C:\\temp\\Schedule.xlsx')
            worksheet = workbook['Schedule']

            p1 = DoughnutChart()
    #Locate data and categories
            labels = Reference(worksheet,
                     min_col=7,
                     #max_col=min_column+8,
                     min_row= 4,
                     max_row= 8) #including headers
            data = Reference(worksheet,
                           min_col=8,
                           #max_col=min_column+6,
                           min_row=3,
                           max_row=min_row+5)
            
    # adding data and categories
            p1.add_data(data, titles_from_data=True)
            p1.set_categories(labels)
            #p1.style = 1 #choose the chart style
            p1.title = 'Breakdown'

    # Cut the first slice out of the pie
            slice = DataPoint(idx=2, explosion=20)
            p1.series[0].data_points = [slice]

    # Showing data labels as percentage
            p1.dataLabels = DataLabelList()
            p1.dataLabels.showPercent = True
            p1.dataLabels.format = '.2f'
            
    #Location chart
            worksheet.add_chart(p1, "F23")
            workbook.save(r'C:\\temp\\Schedule.xlsx')
            return messagebox.showinfo('Schedule.xlsx created', 'Please check c:\ temp folder') 

#Create PDF

        def show_pdf():
        
            if os.path.exists(r'C:\\temp\\Schedule.xlsx'):
                input_file = r'C:\\temp\\Schedule.xlsx'
                #print ("File exist")
            else:
                #print ("File not exist")
                return messagebox.showwarning('Warning', 'Please click "Create Excel Schedule Button"')
           
            input_file = r'C:\\temp\\Schedule.xlsx'
            app = client.DispatchEx("Excel.Application")
            app.Interactive = False
            app.Visible = False
            wb = app.Workbooks.Open(input_file)


        # Create layout
            ws_index_list = [1] #say you want to print these sheets

            #path_to_pdf = r'C:\user\desktop\sample.pdf'
            #give your file name with valid path 
            output_file = r'C:\\temp\\Schedule.pdf'
            
            for index in ws_index_list:

        #Off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

            #ws.PageSetup.PrintArea = print_area

            wb.WorkSheets(ws_index_list).Select()

            wb.ActiveSheet.ExportAsFixedFormat(0, output_file)

            wb.Close()
            app.Quit()
            return messagebox.showinfo('Schedule.pdf created', 'Please check c:\ temp folder')
            

    #Plot (pie chart)

        def plot_show():
    #Message if blank results
        
            if blank_int.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
               blank_int.get()

    # Pie chart, where the slices will be ordered and plotted counter-clockwise:
            labels = 'Starting Amount', 'Yearly Deposit', 'Interest'
            fd = float(prin.get())
            pd = float(pmt.get())
            irr = blank_int.get()

            value = re.sub("[^\d\.]", "", irr)#make numbers only
            ir = float(value)#make float
            #print(ir)

            yr = float(n.get())
            full = fd + (pd*yr) + ir
                   
            #print(full)
            x = (fd/full)*100
            y = ((pd*yr)/full)*100
            z = (ir/full)*100
                   
            sizes = [x,y,z]
            explode = (0, 0, 0.1, )  # only "explode" the 3nd slice (i.e. 'interest')

            fig1, ax1 = plt.subplots()
            ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
                shadow=True, startangle=90)
            ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

            plt.show()
        
    #Go to next entry when "Press Enter" key
        def go_to_next_entry(event, entry_list, this_index):
            next_index = (this_index + 1) % len(entry_list)
            entry_list[next_index].focus_set()

        entries = [child for child in self.winfo_children() if isinstance(child, Entry)]
        for idx, entry in enumerate(entries):
            entry.bind('<Return>', lambda e, idx=idx: go_to_next_entry(e, entries, idx))


        def quit():
            MsgBox = messagebox.askquestion ('Exit Application','Are you sure you want to "Exit"',icon = 'warning')
            if MsgBox == 'yes':
                self.destroy()
            else:
                messagebox.showinfo('Return','You will now return to the application')


    #Activation buttons

        Button(self, text='Show Results',bg = 'green',fg='white',font=('calibri',10, 'bold'),command=lambda:[clearfv(),show_answer()]).grid(row=6, column=1,sticky='news')
        Button(self, text='Show Plot', bg = 'green',fg='white',font=('calibri',10, 'bold'),command=plot_show).grid(row=9, column=1,sticky='news')
        Button(self, text='Clear Results',bg='orange',font=('calibri',10, 'bold'), command=clearfv).grid(row=6, column=4,sticky='news')
        Button(self, text='Clear All',bg='orange',font=('calibri',10, 'bold'), command=clearall).grid(row=7, column=4,sticky='news')
        Button(self, text='Create Excel Schedule',bg = 'green',fg='white',font=('calibri',10, 'bold'), command= show_schedule).grid(row=7, column=1,sticky='news')
        Button(self, text='Create PDF Schedule',bg = 'green',fg='white',font=('calibri',10, 'bold'), command= show_pdf).grid(row=8, column=1,sticky='news')
        prin.focus_set()
        
        Button(self, text='Back to Menu',fg='white',bg = 'red',font=('calibri',10, 'bold'), command= lambda:controller.show_frame(StartPage)).grid(row=8, column=4,sticky='news')


class NPV(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        Label(self, text = "NPV & IRR").grid(row=0,column=1)            
        Label(self,text = "Initial Investment:").grid(row=1)
        Label(self, text = "Yearly Cash Flows, Seperated by Comma:",wraplength=120, justify="right").grid(row=2)
        Label(self, text = "Discount Rate (%):").grid(row=3)
        Label(self, text = "Calculated NPV:",font=('calibri',10, 'bold')).grid (row=1, column=4)
        Label(self, text = "Calculated IRR:",font=('calibri',10, 'bold')).grid (row=2, column=4)
        Label(self, text = "Developer : Sunil",font=('calibri',10, 'bold', 'italic')).place(x =480, y = 300)

    #Entry box for enter details
        i_inv = tk.Entry(self)
        cfy = tk.Entry(self)
        drate = tk.Entry(self)

        #Entry field location
        i_inv.grid(row=1,column=1)
        cfy.grid(row=2,column=1)
        drate.grid(row=3,column=1)
        #n.grid(row=4, column=1)
        #ant.grid(row=5, column=1)

    #Result output field format

        blank_npv= tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_npv.grid(row=1, column=5)

        blank_irr= tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_irr.grid(row=2, column=5)

    #Quit application

        def exit():
            MsgBox = messagebox.askquestion ('Exit Application','Are you sure you want to "Exit"',icon = 'warning')
            if MsgBox == 'yes':
                self.destroy()
            else:
                messagebox.showinfo('Return','You will now return to the application')

    # Clear results fields
        def clearnpv():
            blank_npv.delete(0, END)
            blank_irr.delete(0, END)

    #Clear all fields
        def clearall():
            blank_npv.delete(0, END)
            blank_irr.delete(0, END)
            i_inv.delete(0, END)
            cfy.delete(0, END)
            drate.delete(0, END)
            
    #Set curser focus on to strating amount
            i_inv.focus_set()


        def show_answer():
    
    #Setting variables, data type from entry
    
    #Initial investment
            cfo_x1 = i_inv.get()
            if i_inv.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Initial Investment"')
            try:
                int(cfo_x1)
            except ValueError:
                try:
                    float(cfo_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            cfo_x = float(cfo_x1)
            #print(cfo_x)

    #Yearly Cash Flow
            cfy_x1 = cfy.get().split(',') #Enter values Separated by Commas
            if cfy.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Yearly Cash Flow"')
            elif not re.match("^[0123456789,.]*$", cfy.get()):
                return messagebox.showwarning('Warning', 'Please correct the Yearly Cash Flow Entry')
            #try:
                #int(cfy_x1)
            #except ValueError:
                #try:
                    #float(cfy_x1)
                #except ValueError:
                    #return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            cfy_x = cfy_x1
            #print(cfy_x)

    #Discount rate
            drate_x1 =drate.get()
            if drate.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Discount Rate"')
            try:
                int(drate_x1)
            except ValueError:
                try:
                    float(drate_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            drate_x = float(drate_x1)
            #print(drate_x)

    #Interst calculation factor
            idrate = drate_x/100
            #print(idrate)

    #NPV output
            sum = 0
            t=0
    
            for cf in cfy_x:
                t = t+1
                pv = float(cf)/((1+idrate)**t)
                #print(pv)
                sum = sum + pv
                #print(sum)           
        
            npv = sum - cfo_x
            #print(npv)

            Ans = ('${:,.2f}'.format(npv))
            blank_npv.insert(3,Ans)

    #IRR output
           
            cfo_x = -1*float(cfo_x)
            tcf = [cfo_x] + cfy_x
            irr = npf.irr(tcf)
            #print(irr)

            Ans = ("{:.2%}".format(irr))
            blank_irr.insert(3,Ans)
            return messagebox.showinfo('Created', 'See Results')

    #Go to next entry when "Press Enter" key
        def go_to_next_entry(event, entry_list, this_index):
            next_index = (this_index + 1) % len(entry_list)
            entry_list[next_index].focus_set()

        entries = [child for child in self.winfo_children() if isinstance(child, Entry)]
        for idx, entry in enumerate(entries):
            entry.bind('<Return>', lambda e, idx=idx: go_to_next_entry(e, entries, idx))

        #Button(self, text='Exit',fg='white',bg = 'red',font=('calibri',10, 'bold'), command=exit).grid(row=7, column=6)
        #Button(self, text='Exit',fg='white',bg = 'red',font=('calibri',10, 'bold'), command=self.quit).grid(row=7, column=6)
        Button(self, text='Show Results',bg = 'green',fg='white',font=('calibri',10, 'bold'),command=lambda:[clearnpv(),show_answer()]).grid(row=6, column=1,sticky='news')
        #Button(self, text='Show Plot', bg = 'green',fg='white',font=('calibri',10, 'bold'),command=plot_show).grid(row=9, column=1,sticky='news')
        Button(self, text='Clear Results',bg='orange',font=('calibri',10, 'bold'), command=clearnpv).grid(row=6, column=5,sticky='news')
        Button(self, text='Clear All',bg='orange',font=('calibri',10, 'bold'), command=clearall).grid(row=7, column=5,sticky='news')
        Button(self, text='Back to Menu',fg='white',bg = 'red',font=('calibri',10, 'bold'), command= lambda:controller.show_frame(StartPage)).grid(row=8, column=5,sticky='news') 
        i_inv.focus_set()
      
class Mortgage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        Label(self, text = "Mortgage Calculator").grid(row=0,column=1)
        Label(self,text = "Home Price:").grid(row=1)
        Label(self, text = "Down Paymnet(%):").grid(row=2)
        Label(self, text = "Annual Interest:").grid(row=3)
        Label(self, text = "No of Years:",justify="right").grid (row=4)
        Label(self, text = " Choose Date:",justify="right").grid(row=5)

        Label(self, text = "Mortgage Payment (Monthly):",font=('calibri',10, 'bold'), wraplength=120, justify = RIGHT).grid (row=1, column=3)
        Label(self, text = "Mortgage Payment (Total):",font=('calibri',10, 'bold'),wraplength=120,justify = RIGHT).grid (row=2, column=3)
        Label(self, text = "Total interest:",font=('calibri',10, 'bold')).grid (row=3, column=3)
        Label(self, text = "Developer : Sunil",font=('calibri',10, 'bold', 'italic')).place(x =480, y = 300)

        cal = Calendar(self,
                   font="Arial 14", selectmode='day',
                   cursor="hand1", year=2021, month=2, day=5)
       

    #Entry box for enter details
        hp = tk.Entry(self)
        dp= tk.Entry(self)
        rate = tk.Entry(self)
        yrs = Entry(self)
        cal = DateEntry(self)
        

    #Entry field location
        hp.grid(row=1,column=1)
        dp.grid(row=2,column=1)
        rate.grid(row=3,column=1)
        yrs.grid(row=4,column=1)
        cal.grid(row=5,column=1)
      

    #Result output field format

        blank_mpm = tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_mpm.grid(row=1, column=4)


        blank_mpt = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_mpt.grid(row=2, column=4)


        blank_tint = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_tint.grid(row=3, column=4)


    #Quit application

        def exit():
            MsgBox = messagebox.askquestion ('Exit Application','Are you sure you want to "Exit"',icon = 'warning')
            if MsgBox == 'yes':
                self.destroy()
            else:
                messagebox.showinfo('Return','You will now return to the application')

    # Clear results fields
        def clearresults():
            blank_mpm.delete(0, END)
            blank_mpt.delete(0, END)
            blank_tint.delete(0, END)

    #Clear all fields
        def clearall():
            blank_mpm.delete(0, END)
            blank_mpt.delete(0, END)
            blank_tint.delete(0, END)
            hp.delete(0, END)
            dp.delete(0, END)
            rate.delete(0, END)
            yrs.delete(0, END)
            #ant.delete(0, END)
              
    #Set curser focus on to strating amount
            hp.focus_set()


#CALCULATIONS

        def show_answer():
    
    #Setting variables, data type from entry
    
    #Home price
            hp_x1 = hp.get()
            if hp.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Home Price"')
            try:
                int(hp_x1)
            except ValueError:
                try:
                    float(hp_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            hp_x =float(hp_x1)

    #Down Payment (%)
            dp_x1 = dp.get()
            if dp.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Down Payment"')
            try:
                int(dp_x1)
            except ValueError:
                try:
                    float(dp_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            dp_x =float(dp_x1) 

    #Interest
            rate_x1 =rate.get()
            if rate.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Annual Interest"')
            try:
                int(rate_x1)
            except ValueError:
                try:
                    float(rate_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            rate_x = float(rate_x1)
        
    #years
            yrs_x1 = yrs.get()
            if yrs.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "No of Years"')
            try:
                int(yrs_x1)
            except ValueError:
                try:
                    float(yrs_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            yrs_x = float(yrs_x1)

    #interest factor and downpayment calculator
            ir = float(rate_x/100)#make to percentage
            #print(ir)
            dpp = float(dp_x/100)#make to percentage
            #print(dpp)
            nhp = hp_x-(dpp * hp_x)#home price after downpaymnet
            #print(nhp)
            n = int(12)#months per year
            ii = (1+ir/n)**(n*yrs_x) # EMI calculation factor 1
            #print(ii)
            it = ((1+ir/n)**(n*yrs_x))-1 #EMI calculation factor 2
            #print(it)
                        
    # Mortgaga monthly payment
            #Payment = P x (r / n) x (1 + r / n)^n(t)] / (1 + r / n)^n(t) - 1
            mpm = (float(nhp) * (ir / n)* ii) / (it)
            #print(mpm)
            Ans1 = ('${:,.2f}'.format(mpm))
            blank_mpm.insert(3,Ans1)
            
    # Total mortgage payment
            mpt = yrs_x *n * mpm
            #print(mpt)
            Ans2 = ('${:,.2f}'.format(mpt))
            blank_mpt.insert(3,Ans2)
            
    # Total interest paid
            tint = mpt - nhp
            #print(tint)
            Ans3 = ('${:,.2f}'.format(tint))
            blank_tint.insert(3,Ans3)
            return messagebox.showinfo('Created', 'See Results')
    #Set curser focus on to strating amount
            hp.focus_set()
        

    #Go to next entry when "Press Enter" key
        def go_to_next_entry(event, entry_list, this_index):
            next_index = (this_index + 1) % len(entry_list)
            entry_list[next_index].focus_set()

        entries = [child for child in self.winfo_children() if isinstance(child, Entry)]
        for idx, entry in enumerate(entries):
            entry.bind('<Return>', lambda e, idx=idx: go_to_next_entry(e, entries, idx))
        
        Button(self, text='Show Results',bg = 'green',fg='white',font=('calibri',10, 'bold'),command=lambda:[clearresults(),show_answer()]).grid(row=8, column=1,sticky='news')       
        Button(self, text='Clear Results',bg='orange',font=('calibri',10, 'bold'), command=clearresults).grid(row=7, column=4,sticky='news')
        Button(self, text='Clear All',bg='orange',font=('calibri',10, 'bold'), command=clearall).grid(row=8, column=4,sticky='news')
        Label(self, text = "Developer : Sunil",font=('calibri',10, 'bold', 'italic')).place(x =480, y = 350)
        #Button(self, text='Exit',fg='white',bg = 'red',font=('calibri',10, 'bold'), command=self.quit).grid(row=7, column=6)
        Button(self, text='Back to Menu',fg='white',bg = 'red',font=('calibri',10, 'bold'), command= lambda:controller.show_frame(StartPage)).grid(row=9, column=4,sticky='news')
        hp.focus_set()
        

#CREATE EXCEL SCHEDULE
        def show_schedule():

    #EMI is empty, show message
            
            if blank_mpm.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
                blank_mpm.get()
            
    #Total payment is empty, show message
            
            if blank_mpt.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
               blank_mpt.get()
            
    #Total interest is empty, show message
            
            if blank_tint.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
                blank_tint.get()
        
    #Show Amortization Shedule
    #Declaring variable again and data type
            hp_x1 = hp.get()#total home loan
            hp_x =float(hp_x1)

    #down payment
            dp_x1 = dp.get()
            dp_x =float(dp_x1)

    #Price after downpayment
            dpp = float(dp_x/100)#make to percentage
            #print(dpp)

            nhp = hp_x-(dpp * hp_x)#home price after downpaymnet
            #print(nhp)
        
    #Annula interest rate
            rate_x1 =rate.get()
            rate_x = float(rate_x1)
        
    #Total number of years
            yrs_x1 = yrs.get()
            yrs_x = float(yrs_x1)
     

    #interest in percentage
            n=12
            ir = float(rate_x/100)
            #print(ir)
            ni = ir/int(12)#months per year
            ii = (1+ir/n)**(n*yrs_x) # EMI calculation factor 1
            #print(ii)
            it = ((1+ir/n)**(n*yrs_x))-1 #EMI calculation factor 2
            #print(it)
    #number of months
            yrs_m = yrs_x * n


    # Mortgaga monthly payment
            #Payment = P x (r / n) x (1 + r / n)^n(t)] / (1 + r / n)^n(t) - 1
            mpm = (float(nhp) * (ir / n)* ii) / (it)
            #print(mpm)
            
    # Total Mortgage payment
            mpt = yrs_x *n * mpm
            #print(mpt)
            
    # Total interest payment
            tint = mpt - nhp
            #print(tint)

            #start_date = (date(2021, 1, 1))
            start_date = cal.get()


#https://medium.com/swlh/simple-mortgage-calculator-with-python-and-excel-b98dede36720
    #Panads testing
            rng = pd.date_range(start_date, periods=yrs_m, freq='MS')
            rng.name = "Payment Date"
            df = pd.DataFrame(index=rng, columns=['Monthly Payment','Beginning Balance','Principal Paid','Interest Paid','Cum_Payment','Cum_Interest','Balance'], dtype='float')
            df.reset_index(inplace=True)
            df.index += 1
            df.index.name = "Period"
            
            #Monthly payment, interest calculations
            df["Monthly Payment"] = -1 * npf.pmt(ni,yrs_m,nhp) 
            df["Principal Paid"] = -1 * npf.ppmt(ni,df.index,yrs_m,nhp)
            df["Interest Paid"] = -1 * npf.ipmt(ni,df.index,yrs_m,nhp)
            df["Cum_Interest"] = df['Interest Paid'].cumsum()
            df["Cum_Payment"] = df['Monthly Payment'].cumsum()
            df = df.round(2)
            
            #Ending balance calculation
            df["Balance"] = 0
            df.loc[1, "Balance"] = nhp - df.loc[1,"Principal Paid"]

            for period in range(2, len(df)+1):
                previous_balance = df.loc[period-1, 'Balance']
                principal_paid = df.loc[period, 'Principal Paid']

                if previous_balance == 0:
                    df.loc[period,['Monthly Payment', 'Beginning Balance','Principal Paid','Interest Paid','Cum_Payment','Cum_Interest', 'Balance']] ==0
                    continue
                elif principal_paid<= previous_balance:
                    df.loc[period, 'Balance'] = previous_balance - principal_paid

            #Beginning balance
            df["Beginning Balance"] = 0
            df.loc[1, "Beginning Balance"] = nhp

            for period in range(2, len(df)+1):
                #ending_balance = df.loc[period-1, 'Beginning Balance']
                previous_balance = df.loc[period-1, 'Balance']
                principal_paid = df.loc[period, 'Principal Paid']

                if previous_balance == 0:
                    df.loc[period,['Monthly Payment', 'Beginning Balance','Principal Paid','Interest Paid','Cum_Payment','Cum_Interest', 'Balance']] ==0
                    continue
                elif principal_paid <= previous_balance:
                    #df.loc[period, 'Beginning Balance'] = previous_balance - principal_paid
                    df.loc[period, 'Beginning Balance'] = previous_balance 

            #print(df.to_string())

            #https://pbpython.com/improve-pandas-excel-output.html


    #Variable for gragh and Chart          
            
            dpc = hp_x*(dp_x /100)
            #print(dpc)
            irs = rate_x/100
            #print(irs)
            
    #Prepare Write in Excel
    
            writer = pd.ExcelWriter(r'C:\\temp\\Amortization Schedule.xlsx',engine='xlsxwriter')
            df.to_excel(writer,index=True, sheet_name='Table')
            
            workbook = writer.book
            worksheet = writer.sheets['Table']
            worksheet.set_zoom(90)
            bold_fmt = workbook.add_format({'bold': True})
            money_fmt1 = workbook.add_format({'num_format': '$#,##0.0','bold':True})
            money_fmt = workbook.add_format({'num_format': '$#,##0.0'})
            total_percent_fmt = workbook.add_format({'align': 'right', 'num_format': '#,##0.0%',
                                 'bold': True, 'bottom':6})
            align_fmt = workbook.add_format({'align': 'right',
                                         'bold': True})
            num4 = workbook.add_format({'num_format': 'mmm, d yyyy hh:mm AM/PM'})

    # Formatted as 'dd/mm/yy'
            date_time = datetime.now()
            worksheet.write_datetime('L1', date_time,num4)

            #worksheet.set_column('B1:I1', 1, align_fmt)
            worksheet.set_column('B:I', 1, money_fmt)
            worksheet.set_column('L2:L6', 1, money_fmt)
            worksheet.set_column('K1:K6', 1)
            
            worksheet.write('K1','Developer: Sunil',bold_fmt)
            worksheet.write('K2','Monthly Payment',bold_fmt)
            worksheet.write('K3','Total Payment',bold_fmt)
            worksheet.write('K4','House Price',bold_fmt)
            worksheet.write('K5','Principal',bold_fmt)
            worksheet.write('K6','Interest',bold_fmt)
            worksheet.write('K7','Down Payment',bold_fmt)
            worksheet.write('K8','Interest Rate',bold_fmt)
            
            worksheet.write('L2',mpm,money_fmt1)
            worksheet.write('L3',mpt,money_fmt1)
            worksheet.write('L4',hp_x,money_fmt1)
            worksheet.write('L5',nhp,money_fmt1)
            worksheet.write('L6',tint,money_fmt1)
            worksheet.write('L7',dpc,money_fmt1)
            worksheet.write('L8',irs,total_percent_fmt)
            writer.close()
            
      #Generating Amortization table  in Excel
            workbook = load_workbook(r'C:\\temp\\Amortization Schedule.xlsx')
            worksheet = workbook['Table']

            for col in worksheet.columns:
                 max_length = 0
                 column = col[0].column_letter # Get the column name
                 for cell in col:
                     try: # Necessary to avoid error on empty cells
                         if len(str(cell.value)) > max_length:
                             max_length = len(str(cell.value))
                     except:
                         pass
                 adjusted_width = (max_length + 2) * 1.1
                 worksheet.column_dimensions[column].width = adjusted_width
           
            workbook.save(r'C:\\temp\\Amortization Schedule.xlsx')
            workbook.close()
            
#LINE CHART IN EXCEL
            workbook = load_workbook(r'C:\\temp\\Amortization Schedule.xlsx')
            worksheet = workbook['Table']

    #Cell references (original spreadsheet) 
            min_column = workbook.active.min_column
            max_column = workbook.active.max_column
            
            min_row = workbook.active.min_row
            max_row = workbook.active.max_row

            c1 = LineChart()
    #locate data and categories
            data = Reference(worksheet,
                     min_col=min_column+6,     
                     max_col=max_column-3,
  
                     min_row=min_row,
                     max_row=max_row) #including headers
            
            categories = Reference(worksheet,
                           min_col=min_column,
                           max_col=min_column,
                           min_row=min_row+1,
                           max_row=max_row) #not including headers
    #Adding data and categories
            c1.add_data(data, titles_from_data=True)
            c1.set_categories(categories)
    #Location chart
            worksheet.add_chart(c1, "K10")
            c1.title = 'Mortgage Amortization Graph'
            c1.y_axis.title = 'Value'
            c1.x_axis.title = 'Period'
            c1.height = 10 # default is 7.5
            c1.width = 18 # default is 15
            c1.style = 26 #choose the chart style

            workbook.save(r'C:\\temp\\Amortization Schedule.xlsx') 
    

#PIE CHART IN EXCEL
            workbook = load_workbook(r'C:\\temp\\Amortization Schedule.xlsx')
            worksheet = workbook['Table']
            p1 = DoughnutChart()
    #Locate data and categories
            
            labels = Reference(worksheet,
                     min_col= 11,
                     #max_col=min_column+8,
                     min_row= 5,
                     max_row= 8) #including headers
            data = Reference(worksheet,
                           min_col=12,
                           #max_col=min_column+6,
                           min_row= 4,
                           max_row= min_row +5)
            
    # Adding data and categories
            p1.add_data(data, titles_from_data=True)
            p1.set_categories(labels)
            #p1.style = 1 #choose the chart style
            p1.title = 'Payment Breakdown'
            p1.height = 10 # default is 7.5
            p1.width = 18 # default is 15

    # Cut the first slice out of the pie
            slice = DataPoint(idx=1, explosion=20)
            p1.series[0].data_points = [slice]

    # Showing data labels as percentage
            p1.dataLabels = DataLabelList()
            p1.dataLabels.showPercent = True
            p1.dataLabels.format = '.2f'
            
    #Location chart
            worksheet.add_chart(p1, "K29")
            workbook.save(r'C:\\temp\\Amortization Schedule.xlsx')
            return messagebox.showinfo('Amortization Schedule.xlsx created', 'Please check c:\ temp folder')        
            
        Button(self, text='Create Excel Schedule',bg = 'green',fg='white',font=('calibri',10, 'bold'), command= show_schedule).grid(row=9, column=1,sticky='news')
        hp.focus_set()

#SHOW PLOT

        def plot_show():
    
    #EMI is empty, show message
            
            if blank_mpm.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
                blank_mpm.get()
            
    #Total payment is empty, show message
            
            if blank_mpt.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
               blank_mpt.get()
            
    #Total interest is empty, show message
            
            if blank_tint.get() == "":
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
                blank_tint.get()

    # Pie chart, where the slices will be ordered and plotted counter-clockwise:
            labels = 'Principal', 'Interest'
            home_price = float(hp.get())
            down_payment = float(dp.get())
            irate = float(rate.get())
            yr = float(yrs.get())
            down_payment_dec = float(down_payment/100)#decimal percentage
            
            net_home_price = home_price -(down_payment_dec*home_price)#Principal loan amount after downpayment
            #print(net_home_price)
            total_interest = blank_tint.get()#total interest

            value = re.sub("[^\d\.]", "", total_interest)
            tint = float(value)
            #print(value)

            total_payment = net_home_price + tint #total payment
            #print(total_payment)
                   
            #print(total_payment)
            x = (net_home_price/total_payment)*100
            y = (tint /total_payment)*100
                
            sizes = [x,y]
            explode = (0, 0.1, )  # only "explode" the 2nd slice (i.e. 'interest')

            fig1, ax1 = plt.subplots()
            ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
                shadow=True, startangle=90)
            ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

            plt.show()

        Button(self, text='Show Plot',bg = 'green',fg='white',font=('calibri',10, 'bold'), command= plot_show).grid(row=10, column=1,sticky='news')
        hp.focus_set()

class EffectiveInterest(tk.Frame):
    #https://corporatefinanceinstitute.com/resources/knowledge/finance/effective-annual-interest-rate-ear/

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        #Label(self, text = "FUTURE VALUE").grid(row=0,column=1)
        Label(self,text = "Fixed Deposit:").grid(row=1)
        Label(self, text = "Annual Interest Rate:").grid(row=2)
        Label(self, text = "Number of Years :").grid(row=3)
        Label(self, text = "Annual Effective Rate :",font=('calibri',10, 'bold')).grid (row=0, column=4)
        Label(self, text = "Compounding:",font=('calibri',10, 'bold')).grid (row=0, column=3)
        Label(self, text = "Yearly:",font=('calibri',10, 'bold')).grid (row=1, column=3)
        Label(self, text = "Semi-Annual:",font=('calibri',10, 'bold')).grid (row=2, column=3)
        Label(self, text = "Quarterly:",font=('calibri',10, 'bold')).grid (row=3, column=3)
        Label(self, text = "Monthly:",font=('calibri',10, 'bold')).grid (row=4, column=3)
        Label(self, text = "Daily:",font=('calibri',10, 'bold')).grid (row=5, column=3)
        
        #Label(self, text = "Total Returns",wraplength=90, justify="right").grid (row=0,column=5)
        Label(self, text = "Total Returns:",font=('calibri',10, 'bold')).grid (row=0, column=5)
        Label(self, text = "Total Interest:",font=('calibri',10, 'bold')).grid (row=0, column=6)
        Label(self, text = "Developer : Sunil",font=('calibri',10, 'bold', 'italic')).place(x =480, y = 300)

        Label(self, text = " Choose Date:",justify="right").grid(row=4)

        cal = Calendar(self,font="Arial 14", selectmode='day',
                   cursor="hand1", year=2021, month=2, day=5)

    #Entry box for enter details
        dep = tk.Entry(self)
        rate = tk.Entry(self)
        yrs = tk.Entry(self)
        cal = DateEntry(self)      

        #Entry field location
        dep.grid(row=1,column=1)
        rate.grid(row=2,column=1)
        yrs.grid(row=3, column=1)
        cal.grid(row=4, column=1)


#Effecive rate output box

        blank_yrly = tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_yrly.grid(row=1, column=4)

        blank_syrly = tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_syrly.grid(row=2, column=4)

        blank_qrly = tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_qrly.grid(row=3, column=4)

        blank_mtly = tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_mtly.grid(row=4, column=4)

        blank_dly = tk.Entry(self,font = ('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_dly.grid(row=5, column=4)

 #Total value output box        

        blank_tryl = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_tryl.grid(row=1, column=5)

        blank_trsl = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_trsl.grid(row=2, column=5)

        blank_trql = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_trql.grid(row=3, column=5)

        blank_trml = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_trml.grid(row=4, column=5)
        
        blank_trdl = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_trdl.grid(row=5, column=5) 

  #Total interest output box      

        blank_tntyl = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_tntyl.grid(row=1, column=6)

        blank_tntsl = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_tntsl.grid(row=2, column=6)
        
        blank_tntql = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_tntql.grid(row=3, column=6)
        
        blank_tntml = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_tntml.grid(row=4, column=6)

        blank_tntdl = tk.Entry(self,font=('calibre',10,'bold'),bg = 'grey',fg='white',relief=GROOVE)
        blank_tntdl.grid(row=5, column=6)

    #Quit application

        def quit():
        
            MsgBox = messagebox.askquestion ('Exit Application','Are you sure you want to "Exit"',icon = 'warning')
            if MsgBox == 'yes':
                self.destroy()
            else:
                messagebox.showinfo('Return','You will now return to the application')

    # Clear results fields
        def clearresults():
            blank_yrly.delete(0, END)
            blank_syrly.delete(0, END)
            blank_qrly.delete(0, END)
            blank_mtly.delete(0, END)
            blank_dly.delete(0, END)

            blank_tryl.delete(0, END)
            blank_trsl.delete(0, END)
            blank_trql.delete(0, END)
            blank_trml.delete(0, END)
            blank_trdl.delete(0, END)
            
            blank_tntyl.delete(0, END)
            blank_tntsl.delete(0, END)
            blank_tntql.delete(0, END)
            blank_tntml.delete(0, END)
            blank_tntdl.delete(0, END)

            
    #Focus to strating amount
            #prin.focus_set()

    #Clear all fields
        def clearall():

            blank_yrly.delete(0, END)
            blank_syrly.delete(0, END)
            blank_qrly.delete(0, END)
            blank_mtly.delete(0, END)
            blank_dly.delete(0, END)
            
            blank_tryl.delete(0, END)
            blank_trsl.delete(0, END)
            blank_trql.delete(0, END)
            blank_trml.delete(0, END)
            blank_trdl.delete(0, END)

            
            blank_tntyl.delete(0, END)
            blank_tntsl.delete(0, END)
            blank_tntql.delete(0, END)
            blank_tntml.delete(0, END)
            blank_tntdl.delete(0, END)

            dep.delete(0, END)
            rate.delete(0, END)
            yrs.delete(0, END)
     
    #Set curser focus on to strating amount
            dep.focus_set()

        #CALCULATIONS

        def show_answer(): # not added, for ref only
            def leap_year(year):
    
                if year % 400 == 0:
                    return True
                if year % 100 == 0:
                    return False
                if year % 4 == 0:
                    return True
                return False
            def days_in_month (month,year):
            
                if month in {1, 3, 5, 7, 8, 10, 12}:
                    return 31
                if month == 2:
                    if leap_year(year):
                        return 29
                    return 28
                return 30
    
    #Setting variables, data type from entry
    
    #Deposit
            dep_x1 = dep.get()
            if dep.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Fixed Deposit"')
            try:
                int(dep_x1)
            except ValueError:
                try:
                    float(dep_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            dep_x =float(dep_x1)


    #Interest
            rate_x1 =rate.get()
            if rate.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Annual Interest"')
            try:
                int(rate_x1)
            except ValueError:
                try:
                    float(rate_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            rate_y = float(rate_x1)
            rate_x = rate_y/100

    #number of years
            yrs_x1 =yrs.get()
            if yrs.get() == "":
                return messagebox.showwarning('Warning', 'Please enter "Number of years"')
            try:
                int(yrs_x1)
            except ValueError:
                try:
                    float(yrs_x1)
                except ValueError:
                    return messagebox.showwarning('Warning', 'Please enter "Numbers"')
            yrs_x = float(yrs_x1)
 
    #interest factor calculator
            iryl = float(rate_x)#Yearly
            #print(iryl)
            irsl = float(rate_x/2)#Semi-annual
            #print(irsl)
            irql = float(rate_x/4)#Quarterly
            #print(irql)
            irml = float(rate_x/12)#Monthly
            #print(irml)
            irdl = float(rate_x/365)#Daily
            #print(irdl)

            mkper = int(100)#change to percentage constant
            const = int(1)
            
    #effective interest_Yearly
            #efryl = mkper * (((const + iryl)**yrs_x)-const)
            efryl = mkper * ((const + iryl)-const)
            #print(efryl)
            Ans1 = ('{:,.2f}%'.format(efryl))
            blank_yrly.insert(3,Ans1)
            
    #effective interest_semi yearly            
            #efrsl = mkper * (((const + irsl)**(2*yrs_x))-const)
            efrsl = mkper * (((const + irsl)**(2))-const)
            #print(efrsl)
            Ans2 = ('{:,.2f}%'.format(efrsl))
            blank_syrly.insert(3,Ans2)

    #effective interest_quarterly
            #efrql = mkper * (((const + irql)**(4*yrs_x))-const)
            efrql = mkper * (((const + irql)**(4))-const)
            #print(efrql)
            Ans3 = ('{:,.2f}%'.format(efrql))
            blank_qrly.insert(3,Ans3)

    #effective interest_monthly            
            #efrml = mkper * (((const + irml)**(12*yrs_x))-const)
            efrml = mkper * (((const + irml)**(12))-const)
            #print(efrml)
            Ans4 = ('{:,.2f}%'.format(efrml))
            blank_mtly.insert(3,Ans4)
            
    #effective interest_daily            
            #efrdl = mkper * (((const + irdl)**(365*yrs_x))-const)
            efrdl = mkper * (((const + irdl)**(365))-const)
            #print(efrdl)
            Ans5 = ('{:,.2f}%'.format(efrdl))
            blank_dly.insert(3,Ans5)
                
    # Future value for each compounding types
            fvyl = npf.fv(iryl,yrs_x,0,-dep_x)#yearly
            #print(fvyl)
            Ans1 = ('${:,.2f}'.format(fvyl))
            blank_tryl.insert(3,Ans1)

            fvsl = npf.fv(irsl,yrs_x*2,0,-dep_x)#semi-annualy
            #print(fvsl)
            Ans2 = ('${:,.2f}'.format(fvsl))
            blank_trsl.insert(3,Ans2)
            
            fvql = npf.fv(irql,yrs_x*4,0,-dep_x)#quarterly
            #print(fvql)
            Ans3 = ('${:,.2f}'.format(fvql))
            blank_trql.insert(3,Ans3)  

            fvml = npf.fv(irml,yrs_x*12,0,-dep_x)#monthly
            #print(fvml)
            Ans4 = ('${:,.2f}'.format(fvml))
            blank_trml.insert(3,Ans4) 

            fvdl = npf.fv(irdl,yrs_x*365,0,-dep_x)#daily
            #print(fvdl)
            Ans5 = ('${:,.2f}'.format(fvdl))
            blank_trdl.insert(3,Ans5)

            
    # Total interest gained
            tiyl = fvyl - dep_x #yearly
            Ans1 = ('${:,.2f}'.format(tiyl))
            blank_tntyl.insert(3,Ans1)

            tisl = fvsl - dep_x #semi-yearly
            Ans2 = ('${:,.2f}'.format(tisl))
            blank_tntsl.insert(3,Ans2)

            tiql = fvql - dep_x #quaterly
            Ans3 = ('${:,.2f}'.format(tiql))
            blank_tntql.insert(3,Ans3)

            timl = fvml - dep_x #monthly
            Ans4 = ('${:,.2f}'.format(timl))
            blank_tntml.insert(3,Ans4)
            
            tidl = fvdl - dep_x #daily
            Ans5 = ('${:,.2f}'.format(tidl))
            blank_tntdl.insert(3,Ans5)

            return messagebox.showinfo('Created', 'See Results')
    #Set curser focus on to strating amount
            dep.focus_set()

    # Bar chart

        def bar_chart():

            fvyl = blank_tryl.get()
            
            if blank_tryl.get() == "":
                
                return messagebox.showwarning('Warning', 'Please click "Show Results Button"')
            else:
               blank_tryl.get()

    #Future value
            fvyl_1 = blank_tryl.get()#yearly
            valuey = re.sub("[^\d\.]", "", fvyl_1)#make number only
            fvyl = float(valuey)

            fvsl_1 = blank_trsl.get()#semi-annualy
            values = re.sub("[^\d\.]", "", fvsl_1)
            fvsl = float(values)

            fvql_1 = blank_trql.get() #quarterly
            valueq = re.sub("[^\d\.]", "", fvql_1)
            fvql = float(valueq)
         
            fvml_1 = blank_trml.get()#monthly
            valuem = re.sub("[^\d\.]", "", fvml_1)
            fvml = float(valuem)

            fvdl_1 = blank_trdl.get()#daily
            valued = re.sub("[^\d\.]", "", fvdl_1)
            fvdl = float(valued)

    #Period interest
            tiyl_1 = blank_tntyl.get()#yearly
            ty = re.sub("[^\d\.]", "", tiyl_1)#make number only
            tiyl = float(ty)

            tisl_1 = blank_tntsl.get()#semi-annualy
            ts = re.sub("[^\d\.]", "", tisl_1)
            tisl = float(ts)

            tiql_1 = blank_tntql.get() #quarterly
            tq = re.sub("[^\d\.]", "", tiql_1)
            tiql = float(tq)
         
            timl_1 = blank_tntml.get()#monthly
            tm = re.sub("[^\d\.]", "", timl_1)
            timl = float(tm)

            tidl_1 = blank_tntdl.get()#daily
            td = re.sub("[^\d\.]", "", tidl_1)
            tidl = float(td)

            total_fv = [fvyl, fvsl, fvql, fvml, fvdl]
            total_in = [tiyl, tisl, tiql, timl, tidl]
            labels = ['Yearly', 'Semi-yearly', 'Quarterly', 'Monthly', 'Daily']

#Bar graph
            #x = np.arange(len(labels))  # the label locations
            #width = 0.35  # the width of the bars

            #fig, ax = plt.subplots(1,1, figsize = (13,8))
            #rects1 = ax.bar(x - width/2, total_fv,width, label='Total Value')
            #rects2 = ax.bar(x + width/2, total_in,width, label='Total Interest')
            # Add some text for labels, title and custom x-axis tick labels, etc.
            #ax.set_ylabel('Value')
            #ax.set_title('Total interest & total values with different compounting period')
            #ax.set_xticks(x, labels)
            #ax.legend()
            #ax.bar_label(rects1,fmt='$%.2f', padding=3)
            #ax.bar_label(rects2,fmt='$%.2f', padding=3)
            #ax.bar_label(rects2,fmt='$%.2f', padding=3)
            #fig.tight_layout()
            #plt.show()

#Horizontal bargraph
 
            bar_size = 0.25
            padding = 0.25
            y_locs = np.arange(len(labels)) * (bar_size * 3 + padding)
            #y_locs = np.arange(len(labels))

            fig, ax = plt.subplots(1,1,figsize=(13,8))

            rects1 = ax.barh(y_locs - bar_size, total_fv, align='edge', height=bar_size, color='r', label="Total Value")
            rects2 = ax.barh(y_locs, total_in, align='edge', height=bar_size, color='orange', label="Total Interest")
            #rects3 = ax.barh(y_locs + 2 * bar_size, neutral_vals, align='edge', height=bar_size, color='yellow', label="neutral tweet count")
            ax.set(yticks=y_locs, yticklabels=labels, ylim=[0 - padding, len(y_locs)])

            ax.set_xlabel('Value ($)')
            ax.set_title('Total Interest & Total Values with Different Compounting Period')
            #ax.set_yticks(y_locs,total_fv,total_in )
            ax.legend()

            for y, x in enumerate(total_fv):
                plt.text(x, y, str('${:,.2f}'.format(x)), ha = 'right', va = 'bottom')
            for y, x in enumerate(total_in):
                plt.text(x, y, str('${:,.2f}'.format(x)),ha = 'right',va = 'bottom')

            fig.tight_layout()
            plt.show()
                
      
    #Set curser focus on to fixed deposit
            dep.focus_set()
        

    #Go to next entry when "Press Enter" key
        def go_to_next_entry(event, entry_list, this_index):
            next_index = (this_index + 1) % len(entry_list)
            entry_list[next_index].focus_set()

        entries = [child for child in self.winfo_children() if isinstance(child, Entry)]
        for idx, entry in enumerate(entries):
            entry.bind('<Return>', lambda e, idx=idx: go_to_next_entry(e, entries, idx))
        
        Button(self, text='Show Results',bg = 'green',fg='white',font=('calibri',10, 'bold'),command=lambda:[clearresults(),show_answer()]).grid(row=7, column=1,sticky='news')       
        Button(self, text='Clear Results',bg='orange',font=('calibri',10, 'bold'), command=clearresults).grid(row=7, column=4,sticky='news')
        Button(self, text='Clear All',bg='orange',font=('calibri',10, 'bold'), command=clearall).grid(row=8, column=4,sticky='news')
        Button(self, text='Show Chart',fg='white',bg = 'green',font=('calibri',10, 'bold'), command=bar_chart).grid(row=8, column=1,sticky='news')
        Button(self, text='Back to Menu',fg='white',bg = 'red',font=('calibri',10, 'bold'), command= lambda:controller.show_frame(StartPage)).grid(row=9, column=4,sticky='news')
        dep.focus_set()
        Label(self, text = "Developer : Sunil",font=('calibri',10, 'bold', 'italic')).place(x =480, y = 350)


if __name__ == "__main__":
    app = Financial_Calculator()
    app.mainloop()



    
        







